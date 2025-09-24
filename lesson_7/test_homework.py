import csv
from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from conftest import ZIP_FILE_IN_RESOURCES


def test_csv_in_zip(zip_packing_files):
    with ZipFile(ZIP_FILE_IN_RESOURCES) as zip_file:
        with zip_file.open('example_4kb.csv') as csv_file:
            content = csv_file.read().decode(
                'utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]
            assert second_row[0] == 'Tate Davis'
            assert second_row[1] == '804.613.0262'

def test_pdf_in_zip(zip_packing_files):
    with ZipFile(ZIP_FILE_IN_RESOURCES) as zip_file:
        with zip_file.open('PTWP.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            text = reader.pages[1].extract_text()
            assert 'Python Testing with pytest' in text

def test_xlsx_in_zip(zip_packing_files):
    with ZipFile(ZIP_FILE_IN_RESOURCES) as zip_file:
        with zip_file.open('file_example_XLSX_50.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert 'Mara' == sheet.cell(row=3, column=2).value
